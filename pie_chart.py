import openpyxl
import pandas as pd
import matplotlib.pyplot as plt
# Load the workbook
import openpyxl
import math
import numpy as np
from matplotlib.colors import LinearSegmentedColormap


# 画一些数据（示例）
# ax.plot([1, 2, 3], [4, 5, 6])

# 显示图像
# plt.show()

# 如果需要保存，使用以下代码
# fig.savefig('output.png', dpi=dpi)
# Load the workbook
wb = openpyxl.load_workbook('Geo-AI ethics cases.xlsx')
sheet = wb.active

# Find columns with font size 22 for their header
columns_with_font_size = []
stack=[]
ans=[-1]*sum(1 for _ in sheet.iter_cols(min_row=1, max_row=1))
stack_index=[]
max_value={12:0}
parent_list=[]
child_list=[]
for num,col in enumerate(sheet.iter_cols(min_row=1, max_row=1)):
    cell = col[0]
    columns_with_font_size.append(int(cell.font.sz))
    while stack and stack[-1]<=int(cell.font.sz):
        stack.pop()
        stack_index.pop()
    if stack:
        ans[num]=stack_index[-1]
    stack_index.append(num)
    stack.append(int(cell.font.sz))
    # if int(cell.font.sz)>list(max_value.keys())[0] or int(cell.font.sz)==22 :
    #
    #     max_value[int(cell.font.sz)] = num
    #     # 删除原来的键
    #     if int(cell.font.sz)!=list(max_value.keys())[0]:
    #         del max_value[list(max_value.keys())[0]]
    try:
        # print(int(cell.font.sz),num,stack_index[-2],stack[-2])
        parent_list.append([int(cell.font.sz),num,stack_index[-2],stack[-2]])
    except:
        parent_list.append([int(cell.font.sz), num, -1, -1])
file_path = 'Geo-AI ethics cases.xlsx'
df = pd.read_excel(file_path, engine='openpyxl')
# print(ans)
# print(columns_with_font_size)
# print(parent_list)
parent_node=[100,100]
one_group=[]
def plot_pie_from_index_pairs( index_pairs):
    """
    使用给定的index_pairs在Excel文件中绘制饼图。

    :param file_path: Excel文件的路径。
    :param index_pairs: 嵌套列表，其中每个列表包含两个索引。
    """
    # df = pd.read_excel(file_path, engine='openpyxl')

    n = len(index_pairs)

    # 每行显示的最大饼图数量
    max_cols = 3
    rows = math.ceil(n / max_cols)
    dpi = 100
    figsize = (25.6, 13.29)
    fig, ax = plt.subplots(rows, max_cols, figsize=(40, 7 * rows), dpi=dpi)

    for i, (condition_idx, data_idx) in enumerate(index_pairs):
        row = i // max_cols
        col = i % max_cols

        filtered_data = df[df.iloc[:, condition_idx] == 1].iloc[:, data_idx]
        value_counts = filtered_data.value_counts(normalize=True)
        # print(value_counts)
        ax_pie = ax[row, col] if rows > 1 else ax[col]

        # Get the pie wedges for the legend

        if len(value_counts) == 2:
            # 将标签转换为整数
            colors = ['gray', 'red']

            labels = [str((label)) for label in ["False","True"]]
        else:
            """
            • Level 2 some potential damage
• Level 3 Routine or Significant Potential Hazard
• Level 4 Serious potential hazard
• Level 5 Unsustainable Potential Damage"""

            cmap = LinearSegmentedColormap.from_list("green_cmap", ["pink", "darkred"])
            # 生成四个渐变的绿色
            colors = cmap(np.linspace(0, 1, 4))
            labels = [str(label) for label in ["Level 2\nSome potential damage", "Level 3\nRoutine or Significant Potential Hazard","Level 4\nSerious potential hazard","Level 5\nUnsustainable Potential Damage"]]
        wedges, texts, autotexts = ax_pie.pie(value_counts, colors=colors, autopct='%1.1f%%',
                                              startangle=140)

        ax_pie.legend(wedges, labels, title="Categories", loc="center left", bbox_to_anchor=(1, 0, 0.5, 1),fontsize=14)
        name_a=df.columns[data_idx].replace(".1","").replace("  "," ").lower()
        name_b=df.columns[condition_idx].replace(".1","").replace("  "," ").lower()
        if "fixed" in name_b:
            ax_pie.set_title(f"Proportion of \n {name_a} \n when {name_b}",fontsize=17, fontweight='bold')
            print(f"Proportion of \n {name_a} \n when {name_b}")
        elif "instances" in name_b:
            name_b=name_b.replace("instances with ","")
            ax_pie.set_title(f"Proportion of \n {name_a} \n when {name_b} occur",fontsize=17, fontweight='bold')
            print(f"Proportion of \n {name_a} \n when {name_b} occur")
        else:
            ax_pie.set_title(f"Proportion of \n {name_a} \n when {name_b} exist",fontsize=17, fontweight='bold')
            print(f"Proportion of \n {name_a} \n when {name_b} exist")


        # print(f"Proportion of \n {name_a} \n when {name_b} exist")
        # Calculate total value
        total_value = sum(value_counts)

        # Print proportions for each category
        for label, value in zip(labels, value_counts):
            proportion = (value / total_value) * 100
            print(f"{label}: {proportion:.1f}%")
        print("-------------------")
    # Hide the remaining axes
    for j in range(i + 1, rows * max_cols):
        row = j // max_cols
        col = j % max_cols
        if rows == 1:
            ax[col].axis('off')
        else:
            ax[row, col].axis('off')

    plt.tight_layout()
    print("-------------------")
    dpi = 100


    plt.savefig(r"C:/Users/Morning/Desktop/hiwi/heart/paper/chart/" + str(index_pairs[0][1]+1) + ".png",dpi=dpi)
    # plt.show()
for i in parent_list:
    if i[2]!=-1:
        if i[3]==22 and i[2]!=parent_node[1]:
            parent_node=[22,i[2]]
            # print(parent_node)
            # print(one_group)
            if len(one_group)>=1:
                plot_pie_from_index_pairs(one_group[-1])
                print("-------------------")
            one_group.append([])

            one_group[-1].append([i[2],i[1]])
            try:
                pass
                # print(one_group[-2]) #每个新增的单独的组
            except:
                pass
        else:
            try:

                one_group[-1].append([i[2], i[1]])
            except:
                pass




plot_pie_from_index_pairs(one_group[-1])
# plt.show()