import json
import queue
import random
import time
from infoweb_client import send_message_to_server
import asyncio
import pandas as pd
import openpyxl
# from selenium_chatgpt import selenium_spider
xlsx_path=r'C:\Users\Morning\Desktop\hiwi\heart\paper\file_folder\test_folder\modified_geo_ai_ethics_cases.xlsx'
df = pd.read_excel(xlsx_path)
wb = openpyxl.load_workbook(xlsx_path)
import threading
from gpt_api import change_statement
import re
def has_values_under_header(header_name):

    sheet = wb.active
    header_row = sheet[1]
    # Find the column index for the given header name
    column_index = None
    for cell in header_row:
        if cell.value == header_name:
            column_index = cell.column
            break

    # If header not found, return False
    if column_index is None:
        return False

    # Check cells under the header for values
    for row in range(2, sheet.max_row + 1):  # Start from 2 to skip header row
        cell_value = sheet.cell(row=row, column=column_index).value
        if cell_value is not None and isinstance(cell_value, (int, float)):  # Check if cell has a number
            return True

    return False
def generate_col():


    sheet = wb.active

    # Get the header row
    header_row = sheet[1]

    # Check for bold columns and store the bold column headers
    bold_column_headers = [cell.value for cell in header_row if cell.font.bold]
    # Get indices of the bold columns
    bold_column_indices = [cell.column for cell in header_row if cell.font.bold]

    # Create a dictionary to store headers between bold columns
    corrected_headers_between_bold = {}

    # Iterate over the bold column indices and extract headers between them
    for idx, bold_index in enumerate(bold_column_indices[:-1]):
        next_bold_index = bold_column_indices[idx + 1]
        headers_between = [header_row[i].value for i in range(bold_index, next_bold_index-1)]
        corrected_headers_between_bold[header_row[bold_index-1].value] = headers_between

    print(corrected_headers_between_bold)
    return corrected_headers_between_bold


# for col in corrected_headers_between_bold:
sensitivity="""
personal data revealing racial or ethnic origin, political opinions, religious or philosophical beliefs;
trade-union membership;
genetic data, biometric data processed solely to identify a human being;
health-related data;
data concerning a person’s sex life or sexual orientation.
"""
# if has_values_under_header(col) 后面的键值在键名成立后才考虑
economic_loss={1:'''• A slight economic loss that may equal or slightly exceed daily expenses. • Losses can be recovered quickly without affecting the overall financial condition of the individual or organization.
• For example, loss of small consumer goods or minor investment losses.
''',
                2:'''
Smaller financial losses, but beyond daily budget, may require financial adjustments.
It may take some time to recoup these losses, but there may not be a serious impact on the financial health of the individual or organization in the long run.
'''
                ,3:"""
                • Significant financial losses, which may require significant financial adjustments or changes in strategy.
• For individuals, it may mean losing months of income or significant savings.
• For the organization, it may affect some business operations or cause some projects to fail.
• For example, the failure of a large investment or the loss of an important asset.
                """,
                4:"""
                • Extreme financial losses that may threaten the long-term financial health of an individual or the continued operations of an organization.
• For individuals, it may mean losing several years of income, or a significant portion of life savings.
• For organizations, it may result in mass layoffs, business contraction, or bankruptcy.
• For example, the bankruptcy of a large business or the complete failure of a critical project.
                """,5:"""
                • An absolute economic disaster from which recovery is almost impossible.
• For individuals, it may lead to lifelong poverty or the inability to maintain basic necessities of life.
• For an organization, means outright bankruptcy or dissolution.
• For example, the loss of all an individual's assets, or the complete collapse of a global business.
"""}
physical_harm={1:'''• • Minor scratches or abrasions that do not require medical intervention.
• No ongoing pain or discomfort.


''',
                2:'''
Some minor injuries:
• Superficial cuts, bruises, or minor strains.
• Recovery can be expected within a week.

'''
                ,3:"""
• Medical intervention may be required, such as suturing wounds or fixing fractures.
• Recovery may take a period of time, with potential long-term effects, but full or partial recovery is generally expected.

                """,
                4:"""
• Causes major organ damage or loss of function, such as severe concussion or blindness.
• Requires immediate and ongoing medical intervention.


                """,5:"""
• Injuries that are directly fatal or cause imminent death.
• Causes permanent serious injuries such as high degree of paraplegia.

"""}
discrimination={1:'''• • • Prejudice based on unconscious, non-malicious, everyday minor misunderstandings or beliefs.
• Does not involve any explicit discriminatory behavior or remarks.


''',
                2:'''
• Comments or actions based on stereotypes are, may cause some discomfort but are not sustained or deep.
• The person being discriminated against feels mildly upset or confused.

'''
                ,3:"""
• Long-term, systemic unequal treatment, such as unfair promotion or pay in the workplace.
• People who have been discriminated against may experience severe emotional harm, damaged self-esteem or social isolation.
                """,
                4:"""
• Experiencing physical violence or threats, such as racial or religious hate crimes.
• Seriously affects the quality of life, economic status or personal safety of the person being discriminated against.

                """,5:"""
• Cause serious bodily injury or death.
• Complete deprivation of basic human rights and dignity of those discriminated against, resulting in long-term psychological, social and material deprivation.


"""}
privacy_violation={1:'''
• • Violated information is data that is publicly available or widely known, such as public social media profile information.
• Will not result in any form of discrimination or negative impact.


''',
                2:'''
• The leaked information may be used for harmless commercial purposes, such as advertising.
• May receive unwanted spam or phone calls.


'''
                ,3:"""
• The leaked information could be used for targeted advertising or other more intrusive marketing strategies.
• Be used for demographic profiling that may result in unfair price discrimination or other forms of discrimination.


                """,
                4:"""
• Leaked information, such as health data or biometric data, could be used for discrimination, fraud, or other illegal activities.
• Result in possible financial loss.


                """,5:"""
• Disclosure of information, such as about a person's sexual life, political opinions or religious beliefs, could lead to extreme discrimination, social ostracism or violence.
• A direct threat to an individual's safety, liberty, or quality of life.


"""}
mental_harm={1:'''• • • Mildly unpleasant, probably quickly forgotten.
• Brief mood changes, such as momentary surprise or disappointment.

''',
                2:'''
• Temporary anxiety, irritability, or sadness that does not interfere with daily life.
• is sad about the event, but able to adjust.

'''
                ,3:"""
• Persistent low mood or anxiety that takes some time to recover from.
• Persistent avoidance of certain incident-related situations.

                """,
                4:"""
• Persistent, intense negative emotions such as hopelessness and deep anxiety.
• Significantly affects daily functions such as work, school, or relationships.
                """,5:"""
• Extreme emotional pain, such as deep depression, intense anxiety, or feelings of hopelessness.
• Suicidal thoughts or tendencies toward self-harm may occur.

"""}
def create_data():
    for i in range(40):

        result_dict = generate_col()

        final_dict={}
        question = ""



        for bold_col in result_dict:

            if has_values_under_header(bold_col): #有实际意义的col


                bold_attribute = True if random.random() < 0.5 else False
                final_dict[bold_col] = bold_attribute
                if bold_attribute:
                    if "extent of impact is Identified" in bold_col:
                        question += "%s is True，extent of impact is" % random.choice(result_dict[bold_col])
                    else:
                        question += "%s is True，Detailed rules are as follows:" % bold_col
                        for element in result_dict[bold_col]:

                            if "severity" in element:
                                attribute_value = random.choice([1, 2, 3, 4, 5])
                                question+="%s is \n %s，"%(element, globals().get(str(element).replace("severity of ","").replace(" ",'_').lower())[attribute_value])


                            else:

                                attribute_value = True if random.random() < 0.5 else False
                                question += "%s is %s，"%(element, attribute_value)
                            final_dict[element] = attribute_value


            else:

                question += "\nUnder the topic %s, Detailed rules are as follows:" % bold_col
                for element in result_dict[bold_col]:
                    if "severity" in element:
                        attribute_value = random.choice([1, 2, 3, 4, 5])
                        question += "%s is  \n %s，" % (element, globals().get(
                            str(element).replace("severity of ", "").replace(" ", '_').lower())[attribute_value])


                    else:

                        attribute_value = True if random.random() < 0.5 else False

                        question += "%s is %s，" % (element, attribute_value)
                    final_dict[element] = attribute_value
                    if bold_col=="Data Production Process":
                        question+="Those are all the rules. Please output a news item and do not explicitly display the rules I sent you."
        if i==0:
            selenium_spider("", False, True, "4")
            selenium_spider(
                "I want you to help me generate news about AI accidents in a seroius news style.After I give you specific rules,Please write down the specific accidents that may cause these problems according to the rules I gave, and describe the specific consequences of these accidents in detail, such as what bad consequences were caused and the specific details of these consequences. Details and what are the causes of these consequences, rather than explicitly stating the rules I gave. The output format should be json,like {'News Title':{},'Content':{}}  If you understand, please reply ok")

            result=(selenium_spider(question,True,False,"4"))
        else:
            result=(selenium_spider(question,False,False,"4"))
        print(result)

        print("=========")
        print(final_dict)
        print(result)
        json_dict={"attribute":final_dict,"news":result}
        with open("generated_dict.jsonl", "a", encoding='utf-8') as f:

                f.write(json.dumps(json_dict) + "\n")

def extract_dict(text,question,system_text,key_str="boo"):

    mistake_num=0
    while True:
        mistake_num+=1
        try:
            text = text.lower()
            key_str = key_str.lower()
            match = re.search(r'({[^{}]*})', text)
            json_str = match.group(1) if match else None
            print(json_str)

            if key_str!="multikey":

                result =re.sub(r'[^a-zA-Z]', '', json_str)
                key_str=re.sub(r'[^a-zA-Z]', '', key_str)

                # print(result,key_str)
                if key_str in result:

                    first_letter = result.replace(key_str,"")[0]
                    # print(first_letter)
                    if "f" in first_letter:
                        return False
                    elif "t" in first_letter:
                        return True
                    # print(first_letter)
                else:
                        aa=input("出错了,要继续吗")

                        text=change_statement (question, system_text)
                        continue



            elif key_str=="multikey":
                    json_str = re.sub(r"'", '"', json_str)
                    dictionary = json.loads(json_str)
                    for i in dictionary:
                        if not isinstance(dictionary[i], bool):
                            if "false" in dictionary[i]:
                                dictionary[i]=False
                            elif "true" in dictionary[i]:
                                dictionary[i]=True

                    for i in dictionary:
                        if dictionary[i]==True:

                            return i  #return the key is true
                    print("all false")
                    return list(dictionary.keys())[0]

                    # text = change_statement(question, system_text)
                    # continue

        except Exception as e:
            if mistake_num>=2:
                if key_str=="multikey":
                    return 0
            print("key value error :",e)
            aa = input("出错了,要继续吗")
            text=change_statement (question, system_text)
            print(text)

    # Parse the JSON string to a Python dictionary

def dict_generate(list_):
    dict_={}
    for i in list_:
        dict_["%s"%i]=""
    return dict_
def checking_layer():
    ques_num=0
    description_list=df["description"].fillna("").tolist()

    detailed_description_list=df["detailed description"].fillna("").tolist()
    threads = []
    lock = threading.Lock()
    data_queue = queue.Queue()

    for num, element in enumerate(description_list):
        # print(true_col)
        if num>=4:

            data_queue.put((num,description_list[num] + detailed_description_list[num]))

        # if num>=2:
        #     ques_num+=1
            # selenium_spider("", False, True)

    for i in range(1):
        t = threading.Thread(target=one_process, args=(
            data_queue, lock,i))
        t.start()
        threads.append(t)

    for t in threads:
        t.join()
def one_process(data_queue,lock,index_):
    while True:
        try:
            data_queue_tuple=data_queue.get(timeout=3)
        except:
            break
        print(index_,"processing++++++++++++++++=",data_queue_tuple[0])
        case_str=data_queue_tuple[1]
        system_text = (
                    case_str + ", according to the case above, please answer my question below with json format.{'':ture | false}, and do not output anything else json.")
        print(system_text)
        final_dict = {}
        question = ""
        result_dict = generate_col()
        for bold_col in result_dict:
            # bold_col=bold_col.lower()
            if has_values_under_header(bold_col):  # 有实际意义的col
                question = "Please combine the news above to determine whether %s is true in the news,If true, set the following json value to True, else set following json value to False : %s" % (
                    bold_col, {bold_col: ""})
                boolean_feedback = change_statement(question, system_text)
                # if ques_num==1:
                #     boolean_feedback=selenium_spider(question,True,False)
                # else:
                #     boolean_feedback = selenium_spider(question, False, False)

                bold_attribute = extract_dict(boolean_feedback, question, system_text, bold_col)
                final_dict[bold_col] = bold_attribute
                print({bold_col: bold_attribute})

                if bold_attribute:
                    if "extent of impact is identified" in bold_col.lower():
                        question = "Based on the extend of impact is fixed, Please determine the scope of the incident’s impact in the news,  Set the key you think is correct to True (please set only one key to true):  %s " % dict_generate(
                            result_dict[bold_col])
                        boolean_feedback = change_statement(question, system_text)
                        bold_attribute = extract_dict(boolean_feedback, question, system_text, "multikey")
                        final_dict.update({bold_attribute: True})  # local people:True
                        print({bold_attribute: True})
                    elif "sensitive privacy breach" in bold_col.lower():
                        question = "In terms of sensitive privacy breach, Do you think the following types of privacy data breaches are in the news %s? If true, set the following json value to True, else set following json value to False  :  %s " % (
                        sensitivity, {"sensitive privacy breach": ""})
                        boolean_feedback = change_statement(question, system_text)
                        bold_attribute = extract_dict(boolean_feedback, question, system_text, bold_col)
                        final_dict.update({bold_attribute: True})  # local people:True
                        print({bold_attribute: True})
                    else:

                        for element in result_dict[bold_col]:

                            if "severity" in element.lower():

                                attribute_value = dict_generate([1, 2, 3, 4, 5])
                                question = "In terms of %s ，please determin its severity,according to guideline below: \n%s\n, Please change the key value after the severity you think in the following json to True (least severity is 1)(please only select one severity), and give me this json back%s" % (
                                    bold_col,
                                    globals().get(str(element).replace("severity of ", "").replace(" ", '_')),
                                    attribute_value)
                                boolean_feedback = change_statement(question, system_text)
                                bold_attribute = extract_dict(boolean_feedback, question, system_text, "multikey")
                                # final_dict.update({element:bold_attribute}) # severit2y:5
                            else:
                                question = "Based on %s is True in news, Please combine the news above to determine whether %s is true in the news, If true, set the following json value to True, else set following json value to false: %s" % (
                                bold_col, element, {element: ""})
                                boolean_feedback = change_statement(question, system_text)
                                bold_attribute = extract_dict(boolean_feedback, question, system_text, element)
                            print({element: bold_attribute})
                            final_dict.update({element: bold_attribute})  # severity:5



            else:

                question_first = "Under the topic %s, " % bold_col
                for element in result_dict[bold_col]:
                    if "severity" in element.lower():

                        attribute_value = dict_generate([1, 2, 3, 4, 5])
                        question = question_first + "In terms of %s ，please determin its severity,according to guideline below: \n%s\n, Please change the key value after the severity you think in the following json to True (least severity is 1)(please only select one severity), and give me this json back%s" % (
                            bold_col, globals().get(str(element).replace("severity of ", "").replace(" ", '_')),
                            attribute_value)
                        boolean_feedback = change_statement(question, system_text)

                        bold_attribute = extract_dict(boolean_feedback, question, system_text, "multikey")
                        # final_dict.update({element:bold_attribute}) # severity:5
                    else:
                        question = question_first + " Please combine the news above to determine whether %s is true in the news, If true, set the following json value to True, else set following json value to false: %s" % (
                            element, {element: ""})
                        boolean_feedback = change_statement(question, system_text)
                        bold_attribute = extract_dict(boolean_feedback, question, system_text, element)
                    final_dict.update({element: bold_attribute})  # severity:5
        # json_dict = {"attribute": final_dict, "news": result}
        print(len(final_dict), "+========================================")
        with lock:
            with open("output_final_dict_t.jsonl", "a", encoding='utf-8') as f:
                f.write(json.dumps(final_dict) + "\n")

    # time.sleep(15)
checking_layer()
# ques="""
# {
#   'privacy violation': true
# }"""
# extract_dict(ques,"privacy violation")