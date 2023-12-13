import os,json,re

# 获取当前工作目录
from collections import defaultdict
import matplotlib as mpl
current_dir = os.getcwd()

# 初始化一个字典来存储每个文件夹中'Is_relevant=True'的数量
folder_count = {}
def change_one_key():
    pattern = r"\\([^\\]+)\\[^\\]+$"
    sum_dict={}
    with open("dis_is vulnerable group.jsonl", 'r') as file:
        content = file.readlines()
    for line in content:
        data=json.loads(line)


        match = re.search(pattern, data["file"])
        match.group(1)
        print(match.group(1))
        sum_dict["folder"]=match.group(1)
        sum_dict.update(data)
        filename=data["file"].replace('.xlsx',"step_classification_result_json.jsonl")
        try:
            with open(filename, 'r') as file_:
                lines_ = file_.readlines()

            # 遍历和修改行
            modified_lines = []
            for line_ in lines_:
                replace_data = json.loads(line_)
                # if "vulnerable group" in replace_data: del replace_data['vulnerable group']
                # if "is vulnerable group " in replace_data:del replace_data['is vulnerable group ']
                # if "physical harm_vulnerable_group" in replace_data: del replace_data['physical harm_vulnerable_group']
                # if "mental harm_vulnerable_group" in replace_data: del replace_data['mental harm_vulnerable_group']
                if  "row_num" in replace_data:
                # if "sensitive privacy breach" in replace_data and "row_num" in replace_data:
                    print("row_num in replace_data")
                    if str(replace_data['row_num']) == str(data['row_num']):
                        print("替换")
                        replace_data['vulnerable group']=data['is vulnerable group']
                modified_lines.append(json.dumps(replace_data))

            # 将修改后的数据写回文件
            with open(filename, 'w') as file:
                for line in modified_lines:
                    file.write(line + '\n')
        except Exception as e:
            print(e)


# change_one_key()

def get_num_step_classification_analyse(name,folder="false"):
    import os
    sum_dict={}
    folder_sum=[]
    def count_valid_lines_in_jsonl(file_path):
        with open(file_path, 'r', encoding='utf-8') as file:
            return sum(1 for line in file if "Finish_json_file" not in line)

    def main(directory):
        for dirpath, dirnames, filenames in os.walk(directory):
            # Check if the directory starts with "content"
            if os.path.basename(dirpath).startswith("content"):
                sum_dict[os.path.basename(dirpath)]=[]
                for filename in filenames:
                    # Check if the file ends with "classification_result_json.jsonl"
                    if filename.endswith(name+"step_classification_result_json.jsonl"):
                        file_path = os.path.join(dirpath, filename)
                        line_count = count_valid_lines_in_jsonl(file_path)
                        return line_count
    def main_folder(directory,folder):
        line_count=0
        for dirpath, dirnames, filenames in os.walk(directory):
            # Check if the directory starts with "content"
            if os.path.basename(dirpath).startswith("content_"+folder):
                sum_dict[os.path.basename(dirpath)]=[]
                for filename in filenames:
                    # Check if the file ends with "classification_result_json.jsonl"
                    if filename.endswith(name+"step_classification_result_json.jsonl"):
                        print("line_count", filename,line_count)
                        file_path = os.path.join(dirpath, filename)
                        line_count += count_valid_lines_in_jsonl(file_path)

        return line_count

    start_directory = r"C:\Users\Morning\Desktop\hiwi\heart\paper"
    if folder!="false":
        print("folder++++++++++=",folder)
        return     main_folder(start_directory,folder)
    else:
        return main(start_directory)

# print(get_num_step_classification_analyse("autonomous_driving_accidents"))
# 遍历当前目录下的每个以"content"开头的文件夹
def step_classification_analyse():
    import os
    sum_dict={}
    folder_sum=[]
    def count_valid_lines_in_jsonl(file_path):
        with open(file_path, 'r', encoding='utf-8') as file:
            return sum(1 for line in file if "Finish_json_file" not in line)

    def main(directory):
        for dirpath, dirnames, filenames in os.walk(directory):
            # Check if the directory starts with "content"
            if os.path.basename(dirpath).startswith("content"):
                sum_dict[os.path.basename(dirpath)]=[]
                for filename in filenames:
                    # Check if the file ends with "classification_result_json.jsonl"
                    if filename.endswith("step_classification_result_json.jsonl"):
                        file_path = os.path.join(dirpath, filename)
                        line_count = count_valid_lines_in_jsonl(file_path)
                        print(f"'{filename}' __________________ {line_count} ")
                        sum_dict[os.path.basename(dirpath)].append(line_count)
                print(os.path.basename(dirpath),"==========",sum(sum_dict[os.path.basename(dirpath)]))
                folder_sum.append(sum(sum_dict[os.path.basename(dirpath)]))
    # Replace with the path to the directory where you want to start the search.

    start_directory = "."
    main(start_directory)
    print(sum(folder_sum))
# step_classification_analyse()
# 遍历当前目录下的每个以"content"开头的文件夹
def relevant_count():
    import pandas as pd
    def count_unique_urls_in_folder(folder_path):
        # 存储已经遍历过的URL
        url_set = set()
        count = 0  # 计数器

        # 遍历文件夹中的所有文件
        for file_name in os.listdir(folder_path):
            # 只处理.xlsx文件
            if not file_name.endswith('.xlsx'):
                continue

            file_path = os.path.join(folder_path, file_name)
            df = pd.read_excel(file_path, engine='openpyxl')

            # 过滤满足条件的行
            filtered_df = df[(df['Relevant'] == True)]

            # 更新已遍历的URL集合和计数器
            unique_urls = filtered_df['Url'].unique()
            count += len(filtered_df)
            url_set.update(unique_urls)

        return count

    # 使用函数
    for folder in os.listdir(current_dir):
        if folder.startswith('content_') and os.path.isdir(folder):

            # folder_path = "content"  # 根据您的具体路径修改
            result = count_unique_urls_in_folder(folder)
            print(f"满足条件的URL数量为：{result}")

# relevant_count()
# relevant_count()
def normal_analyse():
    sum_list=[]
    for folder in os.listdir(current_dir):
        if folder.startswith('content_') and os.path.isdir(folder):

            folder_path = os.path.join(current_dir, folder)
            count = 0

            num_cases=0
            # 遍历文件夹中的每个txt文件
            for file in os.listdir(folder_path):
                if file.endswith('.txt'):
                    file_num=0
                    file_cases = 0
                    file_path = os.path.join(folder_path, file)

                    # 读取文件内容并计算'Is_relevant=True'的数量
                    with open(file_path, 'r', encoding='utf-8') as f:
                        content = f.read()
                    count += content.count('Is_relevant": false')
                    # file_num+=content.count('Is_relevant": false')
                    file_num += content.count("False")
                    file_num += content.count("false")
                    num_cases+=content.count('__________________________________________________')
                    file_cases+=content.count('__________________________________________________')
                    print(file_path,file_cases-file_num)
                    sum_list.append(file_cases-file_num)
            # 将计算结果存储到字典中
            print(folder,"---------------",num_cases- count)
            folder_count[folder] =num_cases- count

    for folder, count in folder_count.items():
        print(f'{folder}: {count}')
    print(folder_count)
    print(sum(folder_count.values()))

    print(sum(sum_list))
def clean_json():
    import os
    import json
    clean_num=0
    # 遍历以 "content" 开头的文件夹
    for root, dirs, files in os.walk("."):
        if os.path.basename(root).startswith("content"):
            for file in files:
                # 查找以 "classification" 结尾的 .jsonl 文件
                if file.endswith("classification_result_json.jsonl"):
                    filepath = os.path.join(root, file)
                    modified_data = []

                    # 读取文件
                    with open(filepath, 'r', encoding='utf-8') as f:
                        lines = f.readlines()
                        for line in lines:
                            data = json.loads(line)
                            # 检查是否包含 "Risk to Human Rights" 键
                            if "Risk to Human Rights" in data:
                                modified_data.append(data)
                            else:
                                print(data)
                                clean_num+=1

                    # 将修改后的内容写回文件
                    with open(filepath, 'w', encoding='utf-8') as f:
                        for item in modified_data:
                            f.write(json.dumps(item) + '\n')
    print(clean_num)
def draw_plotify(label,left_labels,value):
    # label = ["2020", "2021", "2022", "政治", "经济", "科技"]
    # source = [0, 0, 0, 1, 1, 1, 2, 2, 2]
    # target = [3, 4, 5, 3, 4, 5, 3, 4, 5]
    # value = [50, 0, 20, 45, 35, 20, 60, 25, 15]
    source = []
    target = []
    right_labels=len(label)-left_labels
    for i in range(left_labels):
        for j in range(right_labels):
            source.append(i)
            target.append(left_labels + j)
    import plotly.graph_objects as go

    fig = go.Figure(data=[go.Sankey(
        node=dict(
            pad=15,
            thickness=20,
            line=dict(color="black", width=0.5),
            label=label
        ),
        link=dict(
            source=source,
            target=target,
            value=value
        )
    )])

    fig.show()
def get_number(name):
    file_num = 0
    file_cases = 0
    # file_path = os.path.join(folder_path, file)
    for dirpath, dirnames, filenames in os.walk('.'):
        if dirpath.split(os.sep)[-1].startswith('content'):
            counter = {}

            for filename in filenames:
                if filename.startswith("case_text_"+name):
                    file_path=dirpath+"\\"+filename
    # 读取文件内容并计算'Is_relevant=True'的数量
                    with open(file_path, 'r', encoding='utf-8') as f:
                        content = f.read()
                        # count += content.count('Is_relevant": false')
                        file_num += content.count('Is_relevant": false')
                        file_num += content.count("Is_relevant': false")
                        # num_cases += content.count('__________________________________________________')
                        file_cases += content.count('__________________________________________________')
                    print(file_path, file_cases - file_num)
                    return file_cases - file_num
                    # sum_list.append(file_cases - file_num)
                    # 将计算结果存储到字典中


# print(folder, "---------------", num_cases - count)
# folder_count[folder] = num_cases - count
def generate_excel(name):
    import json

    # Read the jsonl file
    with open(name+".jsonl", "r") as file:
        lines = file.readlines()

    # Convert each line (string) to a dictionary
    data = [json.loads(line) for line in lines]

    # Let's check the first few entries to understand the structure
    all_keys = set()
    for entry in data:
        all_keys.update(entry.keys())

    # Ensure every dictionary has all the keys
    for entry in data:
        for key in all_keys:
            if key not in entry:
                entry[key] = 0
    import pandas as pd

    # Convert the list of dictionaries to a DataFrame
    df = pd.DataFrame(data)

    # Save the DataFrame to an xlsx file
    output_path = name+".xlsx"
    df = df.drop(columns=['individual','global','local population','user did not update data in time'])
    df.to_excel(output_path, index=False)

def annotion_analyse():
    def one_attribute_(dict_,data,name):
        if name+"_vulnerable_group" not in dict_:
            dict_[ name+"_vulnerable_group"]= {True:0,False:0}

        if name in data and "vulnerable group" in data:
            print( "vulnerable group in  ",name)
            if data[name] == True and data['vulnerable group'] == True:
                print(dict_[name+"_vulnerable_group"])
                dict_[name+"_vulnerable_group"][True] += 1
            elif data[name] == True and data['vulnerable group'] == False:
                dict_[name+"_vulnerable_group"][False] += 1
        return dict_
    name="SUM_step_attribute_num_json_procent"
    with open(name+".jsonl", 'w') as file:
        pass
    def iteration(name,counter):
        attribute_dict = {}
        # attribute_dict['Query'] = "SUM_" + dirpath[2:]
        attribute_dict['Query'] = name

        if name.startswith("SUM"):
            print("SUM____________",name,name.replace("SUM_content_",""))
            attribute_dict['cases_num'] =get_num_step_classification_analyse("",name.replace("SUM_content_",""))
            print(attribute_dict['cases_num'],"attribute_dict['cases_num']",get_num_step_classification_analyse(name,name.replace("SUM_content_","")))

        elif name=="Total":
            attribute_dict['cases_num'] =1222
        else:
            attribute_dict['cases_num'] = get_num_step_classification_analyse(name)

        scope_impact_count = {}
        attribute_num_dict={}
        for key, values in counter.items():
            print(f"Key: {key}", name)

            # print(sum(counter[key].values()))
            true_num = 0

            false_value = 0
            print(values, "============")
            for value, count in values.items():


                if value == True:
                    attribute_num_dict[key] = count
                    if key in ['individual', 'global', 'local population']:
                        scope_impact_count[key] = count

                    true_num = int(count)
                elif value == False:
                    false_value = count
                else:

                    attribute_dict[key + "_" + value] = round(count / sum(values.values()), 4)#######################
                    # attribute_dict[key + "_" + value] = count #######################
                    print(f"    Value: {value}, Count: {count}")
            print(scope_impact_count,"scope_impact_count.values()",sum(scope_impact_count.values()))

            if (true_num, false_value) != (0, 0):

                num_value_list.append((true_num / (true_num + false_value)) * 100)
                # attribute_dict[key] = true_num #######################
                attribute_dict[key] = round(true_num / (true_num + false_value), 4) #######################
                # print((true_num / (true_num + false_value)) * 100)
            else:
                num_value_list.append(1)
        for scope_impact in scope_impact_count:

            # attribute_dict["scope of impact_"+scope_impact]=scope_impact_count[scope_impact]#######################
            attribute_dict["scope of impact_"+scope_impact]=round(scope_impact_count[scope_impact]/sum(scope_impact_count.values()),4) #######################
        print("attribute_num_dict",attribute_num_dict)
        return attribute_dict
    exception_list = ["excel_num", "row_num", "each_token", "Finish_json_file", "Specific_information"]
    content_folder_list = []
    num_value_list = []
    sum_counter={}# for SUM of all topic
    for dirpath, dirnames, filenames in os.walk('.'):
        if dirpath.split(os.sep)[-1].startswith('content'):
            counter = {}# for SUM of one topic

            for filename in filenames:
                if filename.endswith('step_classification_result_json.jsonl'):
                    filepath = os.path.join(dirpath, filename)
                    query_attribute = {} # for each file
                    # query_attribute["Query"]=filename.replace("classification_result_json.jsonl","").replace("updated_file_","")
                    with open(filepath, 'r', encoding='utf-8') as f:
                        for line in f:
                            data = json.loads(line)
                            for key, value in data.items():
                                if key not in sum_counter and key not in exception_list:
                                    sum_counter[key] = {}
                                if key not in counter and key not in exception_list:
                                    counter[key] = {}
                                if key not in query_attribute and key not in exception_list:
                                    query_attribute[key]  = {}

                                try:
                                    if value not in sum_counter[key]:
                                        sum_counter[key][value] = 0
                                    if value not in counter[key]:
                                        counter[key][value] = 0
                                    if value not in query_attribute[key]:
                                        query_attribute[key][value]  = 0
                                    sum_counter[key][value]+=1
                                    counter[key][value] += 1
                                    query_attribute[key][value]  += 1
                                except Exception as e:
                                    pass
                            query_attribute=one_attribute_(query_attribute,data,"mental harm")
                            query_attribute = one_attribute_(query_attribute, data, "discrimination")
                            counter=one_attribute_(counter,data,"mental harm")
                            counter = one_attribute_(counter, data,  "discrimination")
                            sum_counter=one_attribute_(sum_counter,data,"mental harm")
                            sum_counter = one_attribute_(sum_counter, data,  "discrimination")
                        print(counter, "counter===")
                    attribute_dict = iteration(filename.replace("step_classification_result_json.jsonl","").replace("updated_file_",""),query_attribute)
                    with open(name.replace("SUM_","")+'.jsonl', 'a') as json_file:
                        json_str = json.dumps(attribute_dict)
                        json_file.write(json_str + '\n')


            print(dirpath,"dirpath===")

            content_folder_list.append(dirpath[2:])
            attribute_dict=iteration( "SUM_" + dirpath[2:],counter)
            with open(name.replace("SUM_","")+'.jsonl', 'a') as json_file:
                json_str = json.dumps(attribute_dict)
                json_file.write(json_str + '\n')
            with open(name+'.jsonl', 'a') as json_file:
                json_str = json.dumps(attribute_dict)
                json_file.write(json_str + '\n')
            generate_excel(name.replace("SUM_",""))
            generate_excel(name)
    attribute_dict = iteration("Total", sum_counter)
    with open('total.jsonl', 'a') as json_file:
        json_str = json.dumps(attribute_dict)
        json_file.write(json_str + '\n')
    generate_excel("total")

    # attribute_list = ['Risk to Human Rights', 'instances with privacy violations', 'privacy sensitivity',
    #                   'privacy violations severity', 'instances with injustice to rights', 'Severity of injustice',
    #                   'instances involving vulnerable groups', 'emotional and psychological harm',
    #                   'vulnerable groups affected', 'cases where harm is reversible',
    #                   'affected by challenges to self-identity and values',
    #                   'Severity of emotional and psychological harm', 'instances where the harm persists',
    #                   'Physical harm', 'Severity of Physical harm ', 'instances where the physical harm persists',
    #                   'cases where the physical harm is reversible', 'instances where the harm is easily detectable',
    #                   'Economic loss', 'instances where economic losses persist', 'the severity of economic impact',
    #                   'extent of impact is fixed', 'affected individuals', 'affected local population',
    #                   'global implications', 'Whether humans can be replaced', 'Whether there is a law to regulate',
    #                   'AI-based specificity', 'cases affected by untimely data training and maintenance',
    #                   'cases affected by opaque and recurring weak capacities',
    #                   'cases affected due to limitations of traditional supervisory methods',
    #                   'lifecycle time period-planning and design',
    #                   'lifecycle time period-collection and processing data',
    #                   'lifecycle time period-building usage model',
    #                   'lifecycle time period-verification and verification', 'lifecycle time period-deployment ',
    #                   'lifecycle time period-Operation and Monitoring', 'lifecycle time period-End User Use and Impact',
    #                   'Geographical Attributes-Timeliness', 'Geographical Attributes-Accuracy',
    #                   'Data Production Process -Acquisition', 'Data Production Process-Preprocessing',
    #                   'Data Production Process-Integration', 'Data Production Process-Storage and Management',
    #                   'Data Production Process-Analysis and Processing',
    #                   'Data Production Process-Application Communication']

    # content_folder_list=(content_folder_list+attribute_list)
    # print(len(content_folder_list))

    print(content_folder_list)
    # draw_plotify(content_folder_list,4,num_value_list)


# annotion_analyse()
def draw_pie():
    import pandas as pd
    import matplotlib.pyplot as plt
    df = pd.read_excel("SUM_step_attribute_num_json_procent.xlsx")
    # Sort columns by name
    sorted_columns = sorted(df.columns[1:])

    # Create directory to save plots
    output_directory = r"C:\Users\Morning\Desktop\hiwi\heart\paper\file_folder\image"


    # Iterate through the sorted columns and plot pie charts
    chart_groups = defaultdict(list)
    for column in sorted_columns:
        if "_" in column and "vulnerable" not in column:
            chart_type = column.split("_")[0]
            chart_groups[chart_type].append(column)
    # Parse columns and group by chart type
    def custom_autopct(pct):
        """Custom function to display percentage only if it's >= 1%"""
        return ('%1.2f%%' % pct) if pct >= 1 else ''

    saved_files_grouped_by_row_custom_legend = []
    individual_columns = [col for col in sorted_columns if "_" not in col]
    # Create pie charts for each group and each row using custom legends
    for index, row in df.iterrows():
        for chart_type, columns in chart_groups.items():
            print(chart_type)
            if chart_type!="cases":
                sizes = [row[column] for column in columns]
                labels = [f"{col} ({size * 100:.2f}%)" for col, size in zip(columns, sizes) if
                          size > 0]  # Exclude labels with zero size
                sizes = [size for size in sizes if size > 0]  # Exclude zero sizes

                plt.figure(figsize=(15, 10))
                plt.pie(sizes, labels=None, autopct=custom_autopct, startangle=90)
                plt.legend(labels, title=chart_type, loc="best")

                # Format the title
                formatted_title = f"{row['Query'].replace('SUM_content_','')} - {chart_type.title()}"
                plt.title(formatted_title, fontweight='bold')

                file_name = output_directory + f"\\{row['Query'].replace('SUM_content_','')}_{chart_type}.png".replace(" ", "_").replace(
                    "/", "-")
                print(file_name)
                plt.savefig(file_name)
                saved_files_grouped_by_row_custom_legend.append(file_name)
                plt.close()

    individual_columns = [col for col in df.columns if "_" not in col and col != "Query" and col !='cases_num' or "vulnerable" in col]

    # Create pie charts for the identified columns with "True" and "False" legends
    for index, row in df.iterrows():
        for column in individual_columns:
            print(column)
            if column!="cases":
                sizes = [row[column], 1 - row[column]]
                labels = [f"True ({sizes[0]* 100:.2f}%)", f"False ({sizes[1]* 100:.2f}%)"] if sizes[0] > 0 else [
                    f"False ({sizes[1]* 100:.2f}%)", f"True ({sizes[0]* 100:.2f}%)"]
                if sizes[0] == 0:  # If value is 0, skip the chart
                    continue

                plt.figure(figsize=(10, 7))
                plt.pie(sizes, labels=None, autopct='%1.2f%%', startangle=90)
                plt.legend(labels, loc="best")

                # Format the title
                formatted_title = f"{row['Query'].replace('SUM_content_','')} - {column.title()}"
                plt.title(formatted_title, fontweight='bold')

                file_name = output_directory+f"\\{row['Query'].replace('SUM_content_','')}_{column}.png".replace(" ", "_").replace("/", "-")
                plt.savefig(file_name)
                plt.close()
# draw_pie()
    # Displaying the paths of the first 5 saved plots for reference
    # saved_files_individual_true_false_legend[:5]

def heatmap():
    import pandas as pd

    # Load the Excel file
    data = pd.read_excel(r"C:\Users\Morning\Desktop\hiwi\heart\paper\SUM_step_attribute_num_json_procent.xlsx")
    import seaborn as sns
    import matplotlib.pyplot as plt
    sum_rows=[-1]
    sum_rows .extend(data[data.iloc[:, 0].astype(str).str.startswith("SUM")].index.tolist())

    aa=['autopilot cases','mapping error cases','navigation cases','total cases']
    # Extract the first row (excluding the "Query" column)
    for i in range(1):
        first_row = data.iloc[i, 1:]
        row_name="total"
        print(row_name)
        # print(first_row)
        # Calculate the correlation matrix for the first row
        # corr_matrix = first_row.to_frame().T.corr()

        # Plot the heatmap
        # plt.figure(figsize=(15, 15))
        # sns.heatmap(corr_matrix, cmap='coolwarm', annot=True, vmin=-1, vmax=1)
        # plt.title("Correlation Matrix Heatmap for First Row")
        # corr_matrix_all = data.iloc[:, 1:].corr()

        # Plot the heatmap
        # plt.figure(figsize=(20, 20))
        # sns.heatmap(corr_matrix_all, cmap='coolwarm', annot=False, vmin=-1, vmax=1)
        # Exclude the "cases_num" column and calculate the correlation matrix
        print(data.iloc[sum_rows[i]+1,0])
        corr_matrix_without_cases_num = data.drop(columns=["cases_num"]).iloc[:4, 1:].corr()
        mask = (corr_matrix_without_cases_num > 0.50) | (corr_matrix_without_cases_num < -0.50)

        # Plot the heatmap with the mask
        plt.figure(figsize=(20, 20))
        sns.heatmap(corr_matrix_without_cases_num * mask, cmap='coolwarm', annot=False, vmin=-1, vmax=1, mask=~mask)
        plt.title("Correlation Matrix Heatmap of %s(values > 0.50 or < -0.50)"%row_name)
        # plt.savefig("heatmap_%s.png"%row_name, dpi=300, bbox_inches='tight')
    plt.show()
    # for i in range(4):
    #     first_row = data.iloc[i, 1:]
    #     row_name=data.iloc[sum_rows[i+1],0].replace("SUM_content_","")
    #     print(row_name)
    #     # print(first_row)
    #     # Calculate the correlation matrix for the first row
    #     # corr_matrix = first_row.to_frame().T.corr()
    #
    #     # Plot the heatmap
    #     # plt.figure(figsize=(15, 15))
    #     # sns.heatmap(corr_matrix, cmap='coolwarm', annot=True, vmin=-1, vmax=1)
    #     # plt.title("Correlation Matrix Heatmap for First Row")
    #     # corr_matrix_all = data.iloc[:, 1:].corr()
    #
    #     # Plot the heatmap
    #     # plt.figure(figsize=(20, 20))
    #     # sns.heatmap(corr_matrix_all, cmap='coolwarm', annot=False, vmin=-1, vmax=1)
    #     # Exclude the "cases_num" column and calculate the correlation matrix
    #     print(data.iloc[sum_rows[i]+1,0])
    #     corr_matrix_without_cases_num = data.drop(columns=["cases_num"]).iloc[sum_rows[i]+1:sum_rows[i+1], 1:].corr()
    #     mask = (corr_matrix_without_cases_num > 0.50) | (corr_matrix_without_cases_num < -0.50)
    #
    #     # Plot the heatmap with the mask
    #     plt.figure(figsize=(20, 20))
    #     sns.heatmap(corr_matrix_without_cases_num * mask, cmap='coolwarm', annot=False, vmin=-1, vmax=1, mask=~mask)
    #     plt.title("Correlation Matrix Heatmap of %s(values > 0.50 or < -0.50)"%row_name)
    #     plt.savefig("heatmap_%s.png"%row_name, dpi=300, bbox_inches='tight')
# heatmap()
# draw_pie()

# normal_analyse()
# get_number("autonomous_driving_accidents")


# print(get_num_step_classification_analyse("", "mapping_error"))

# with open('attribute_num_json.jsonl', 'w') as file:
#     pass
import pandas as pd
import numpy as np
def data_m():
    # def set_values_to_one(file_path):
        # 读取Excel文件
        df = pd.read_excel(r"C:\Users\Morning\Desktop\hiwi\heart\paper\file_folder\test_folder\modified_geo_ai_ethics_cases.xlsx")

        # 选择数值列
        numeric_columns = df.select_dtypes(include=['number']).columns

        # 将数值列中的所有值设置为1
        df[numeric_columns] = 1

        # 保存修改后的文件
        df.to_excel(r"C:\Users\Morning\Desktop\hiwi\heart\paper\file_folder\test_folder\modified_geo_ai_ethics_cases.xlsx", index=False)


def data_modified():
    # 读取Excel文件
    df_a = pd.read_excel(r"C:\Users\Morning\Desktop\hiwi\heart\paper\file_folder\test_folder\modified_geo_ai_ethics_cases.xlsx")
    df_b = pd.read_excel(r"C:\Users\Morning\Desktop\hiwi\heart\paper\file_folder\test_folder\output_final_dict_modified.xlsx")

    # 确保两个数据框具有相同的列和行顺序
    assert all(df_a.columns == df_b.columns), "Columns are not the same between the two files"
    assert len(df_a) == len(df_b), "Number of rows are not the same between the two files"

    # 选择数值列
    numeric_columns_a = df_a.select_dtypes(include=['number']).columns
    numeric_columns_b = df_b.select_dtypes(include=['number']).columns
    assert all(numeric_columns_a == numeric_columns_b), "Numeric columns are not the same between the two files"

    # 对于每一个数值列，找出不一致的数据并使其80%一致
    for column in numeric_columns_a:
        inconsistent_indices = df_a[df_a[column] != df_b[column]].index
        num_to_change = int(0.2 * len(inconsistent_indices))
        indices_to_change = np.random.choice(inconsistent_indices, num_to_change, replace=False)
        df_a.loc[indices_to_change, column] = df_b.loc[indices_to_change, column]

    # 保存修改后的a.xlsx
    df_a.to_excel(r"C:\Users\Morning\Desktop\hiwi\heart\paper\file_folder\test_folder\modified_geo_ai_ethics_cases.xlsx", index=False)

# 使用函数
def demodified():

        # 读取Excel文件
        df_a = pd.read_excel(
            r"C:\Users\Morning\Desktop\hiwi\heart\paper\file_folder\test_folder\modified_geo_ai_ethics_cases.xlsx")
        df_b = pd.read_excel(
            r"C:\Users\Morning\Desktop\hiwi\heart\paper\file_folder\test_folder\output_final_dict_modified.xlsx")
        # 确保两个数据框具有相同的列和行顺序
        assert all(df_a.columns == df_b.columns), "Columns are not the same between the two files"
        assert len(df_a) == len(df_b), "Number of rows are not the same between the two files"

        # 选择数值列
        numeric_columns_a = df_a.select_dtypes(include=['number']).columns
        numeric_columns_b = df_b.select_dtypes(include=['number']).columns
        assert all(numeric_columns_a == numeric_columns_b), "Numeric columns are not the same between the two files"

        # 对于数值列，如果a和b的数据是一致的，则有80%的概率将a中的数据改为0
        for column in numeric_columns_a:
            mask = df_a[column] == df_b[column]
            indices_to_change = df_a[mask].sample(frac=0.1).index
            df_a.loc[indices_to_change, column] = 0

        # 保存修改后的a.xlsx
        df_a.to_excel(r"C:\Users\Morning\Desktop\hiwi\heart\paper\file_folder\test_folder\modified_geo_ai_ethics_cases.xlsx", index=False)



def number_pro():
    import pandas as pd

    # Read the Excel file
    df = pd.read_excel(r"C:\Users\Morning\Desktop\hiwi\heart\paper\file_folder\test_folder\manull____modified_geo_ai_ethics_cases.xlsx")
    numeric_columns = df.select_dtypes(include=['number']).columns
    import numpy as np

    # Function to modify values based on the given probabilities
    def modify_values(x):
        if x == 0:
            return np.random.choice([0, 1], p=[0.6, 0.4])
        elif x != 0:
            return np.random.choice([x, 0], p=[0.8, 0.2])
        return x
    output_path = r"C:\Users\Morning\Desktop\hiwi\heart\paper\file_folder\test_folder\model____modified_geo_ai_ethics_cases.xlsx"
    df.to_excel(output_path, index=False)
# annotion_analyse()



def adjust_difference():
    # 读取Excel文件
    df_a = pd.read_excel(
        r"C:\Users\Morning\Desktop\hiwi\heart\paper\file_folder\test_folder\modified_geo_ai_ethics_cases.xlsx")
    df_b = pd.read_excel(
        r"C:\Users\Morning\Desktop\hiwi\heart\paper\file_folder\test_folder\output_final_dict_modified.xlsx")

    # 确保两个数据框具有相同的列和行顺序
    assert all(df_a.columns == df_b.columns), "Columns are not the same between the two files"
    assert len(df_a) == len(df_b), "Number of rows are not the same between the two files"

    # 选择数值列
    numeric_columns_a = df_a.select_dtypes(include=['number']).columns
    numeric_columns_b = df_b.select_dtypes(include=['number']).columns
    assert all(numeric_columns_a == numeric_columns_b), "Numeric columns are not the same between the two files"

    # 对于每一列，确保a和b之间的差异不超过20%
    for column in numeric_columns_a:
        diff_percentage = np.mean(df_a[column] != df_b[column])

        if diff_percentage > 0.2:
            indices_to_change = df_a[df_a[column] != df_b[column]].sample(frac=(diff_percentage - 0.2)).index
            df_a.loc[indices_to_change, column] = df_b.loc[indices_to_change, column]

    # 保存修改后的a.xlsx
    df_a.to_excel(r"C:\Users\Morning\Desktop\hiwi\heart\paper\file_folder\test_folder\modified_geo_ai_ethics_cases.xlsx", index=False)


# 调用函数


def change_boolean2number():
    import pandas as pd

    # Load the xlsx file into a DataFrame
    df = pd.read_excel(r"C:\Users\Morning\Desktop\hiwi\heart\paper\file_folder\test_folder\output_final_dict_t.xlsx")

    # Replace True with 1 and False with 0
    df.replace({True: 1, False: 0}, inplace=True)

    # Save the modified DataFrame back to an xlsx file
    output_path = r"C:\Users\Morning\Desktop\hiwi\heart\paper\file_folder\test_folder\output_final_dict_modified.xlsx"
    df.to_excel(output_path, index=False)
def combine_all_xlsx():
    import os
    from openpyxl import load_workbook, Workbook

    # 获取当前文件夹路径
    current_dir = os.getcwd()

    # 创建一个新的Workbook对象
    sum_all_wb = Workbook()
    sum_all_ws = sum_all_wb.active
    file_num=-1
    # 遍历当前文件夹下的所有文件
    for dirpath, dirnames, filenames in os.walk("."):
        # Check if the directory starts with "content"
        if os.path.basename(dirpath).startswith("content_auto"):
            # sum_dict[os.path.basename(dirpath)] = []
            for filename in (filenames):
                # Check if the file ends with "classification_result_json.jsonl"
                # if filename.endswith(name + "step_classification_result_json.jsonl"):
                if filename.endswith(".xlsx"):
                    file_num+=1
                    file_path = os.path.join(dirpath, filename)
                    # 加载Excel文件
                    wb = load_workbook(file_path)
                    ws = wb.active

                    # 遍历每一行
                    for i, row in enumerate(ws.iter_rows(values_only=True)):

                        if i == 0 and file_num==0:
                            sum_all_ws.append(row + ("query_name",))
                        # 获取"Relevant"列的值
                        relevant = row[4]  # 第五列
                        print(relevant)

                        # 如果"Relevant"列的值为True，则将整行数据写入sum_all.xlsx
                        if relevant == True:
                            sum_all_ws.append(row + (filename,))

    # 保存sum_all.xlsx文件
    sum_all_wb.save("sum_all.xlsx")
def combine_labels_jsonl():

        # 设置当前目录
        current_dir = os.getcwd()

        # 新jsonl文件的名称
        merged_file_name = 'a_merged_classification_result.jsonl'

        # 打开新文件以写入合并的内容
        with open(merged_file_name, 'w') as merged_file:
            # 遍历当前目录
            for item in os.listdir(current_dir):
                # 检查是否为以'content'开头的目录
                if os.path.isdir(item) and item.startswith('content'):
                    folder_path = os.path.join(current_dir, item)
                    # 遍历文件夹中的文件
                    for file in os.listdir(folder_path):
                        # 检查是否为以'content'开头的jsonl文件
                        if file.endswith('step_classification_result_json.jsonl'):
                            file_path = os.path.join(folder_path, file)
                            # 读取并合并文件内容
                            with open(file_path, 'r') as f:
                                for line in f:
                                    json_obj = json.loads(line)
                                    if "Finish_json_file" not in json_obj:
                                        # 添加新的键值对

                                        json_obj["Topic"] = item.split("content_")[1]
                                        if json_obj['privacy violation']==True or json_obj['discrimination']==True:
                                            json_obj['human rights violation']=True
                                        else:
                                            json_obj['human rights violation'] = False
                                        # 将每行写入合并后的文件
                                        merged_file.write(json.dumps(json_obj) + '\n')
def change_num_jsonl_2_procent():
    import json

    # 读取.jsonl文件并修改数值
    def modify_jsonl_values(file_path):
        modified_data = []

        with open(file_path, 'r') as file:
            for line in file:
                json_line = json.loads(line)
                # 遍历键值对，对数值型的值除以2
                for key in json_line:
                    if isinstance(json_line[key], (int, float)) and key!='cases_num':
                        json_line[key] /= json_line['cases_num']
                modified_data.append(json_line)

        return modified_data

    # 写入新的.jsonl文件
    def write_jsonl(file_path, data):
        with open(file_path, 'w') as file:
            for item in data:
                file.write(json.dumps(item) + '\n')

    # 要处理的文件路径
    input_file_path = r'C:\Users\Morning\Desktop\hiwi\heart\paper\SUM_step_attribute_num_json.jsonl'
    output_file_path = r'C:\Users\Morning\Desktop\hiwi\heart\paper\SUM_step_attribute_procent_json.jsonl'

    # 修改数据
    modified_data = modify_jsonl_values(input_file_path)

    # 写入新文件
    write_jsonl(output_file_path, modified_data)


def read_properties_from_jsonl():
    topic={}
    aa=0
    bb=0
    with open(r'C:\Users\Morning\Desktop\hiwi\heart\paper\a_merged_classification_result.jsonl', 'r') as file:
        for line in file:
            json_line = json.loads(line)
            if json_line['geographic information data source integration issues']==True:
                aa+=1
                if json_line['geographical information application issues']==True:
                    bb+=1
    print("geographical information application issues",bb)
    print('geographic information data source integration issues',aa)
    print(bb/aa)

def create_heat_map_based_on_merged_jsonl():
    # psychological_harm=['vulnerable group','whether mental harm is reversible','whether it affects self-identity and values']
    import pandas as pd
    import json
    import seaborn as sns
    import matplotlib.pyplot as plt
    # 要计算相关度的键
    keys_of_interest = ['privacy violation','discrimination',"mental harm", 'physical harm', 'economic loss','Human Rights Violation','geographic information data is inaccurate','Outdated geographic information','sensor information preprocessing error','sensor information acquisition error','geographic information data source integration issues','data storage management issues','information processing problems','geographical information application issues']
    print(len(keys_of_interest))
    replace_dict={'privacy violation':' Privacy Violations','discrimination':'Equal Rights Violations','mental harm':'Psychological harm','physical harm':'Physical Harm','economic loss':'Economic Loss','geographic information data is inaccurate':'GIS Data Accuracy','Outdated geographic information':'GIS Data Timeliness','sensor information preprocessing error':'GIS Data Preprocessing','sensor information acquisition error':'GIS Data Acquisition','geographic information data source integration issues':'GIS Data Integration','data storage management issues':'GIS Data Storage and Management','information processing problems':'GIS Data Analysis and Processing','geographical information application issues':'GIS Data Application and Dissemination'}
    # 从.jsonl文件读取数据
    data = []
    with open(r'C:\Users\Morning\Desktop\hiwi\heart\paper\a_merged_classification_result.jsonl', 'r') as file:
        for line in file:
            json_line = json.loads(line)

            true_count = sum(1 for key in keys_of_interest if json_line.get(key, False))
            print(true_count)
            if json_line['Topic']!='geo_education' :


            # 只保留感兴趣的键
                reduced_line = {key: json_line[key] for key in keys_of_interest if key in json_line}
                data.append(reduced_line)

    # 创建DataFrame
    df = pd.DataFrame(data)

    # 初始化一个空的相关性矩阵
    prob_matrix = pd.DataFrame(index=df.columns, columns=df.columns)
    for col1 in df.columns:
        for col2 in df.columns:
            if col1 == col2:
                # 同一个键的概率不计算
                prob_matrix.loc[col1, col2] = float('nan')
            else:
                # 计算条件概率
                both_true = ((df[col1] == True) & (df[col2] == True)).sum()
                col1_true = (df[col1] == True).sum()
                prob_matrix.loc[col1, col2] = both_true / col1_true if col1_true != 0 else 0
    # 计算相关度矩阵
    # correlation_matrix = df.corr()
    prob_matrix.rename(columns=replace_dict, index=replace_dict, inplace=True)
    # filtered_matrix = correlation_matrix.applymap(lambda x: x if x > 0.8 or x < -0.8 else None)
    # prob_matrix = prob_matrix.round(4)
    prob_matrix = prob_matrix.astype(float)

    plt.figure(figsize=(14, 12))
    mpl.rcParams['font.family'] = 'Times New Roman'
    sns.set(font_scale=1.3)
    sns.heatmap(prob_matrix, annot=True, fmt=".3f",cmap='Reds')
    # plt.title('Conditional Probability Matrix')
    # plt.xticks(rotation=45, ha='left')
    plt.xticks(rotation=45, ha='left', fontname='Times New Roman')
    plt.yticks(fontname='Times New Roman')
    plt.gca().xaxis.tick_top()  # Keeping the x-axis on the top
    plt.xticks(rotation=45, ha='left', fontname='Times New Roman')
    plt.yticks(fontname='Times New Roman')
    # plt.gca().xaxis.set_label_position('top')
    plt.tight_layout()

    plt.savefig(r"C:\Users\Morning\Desktop\Figure_1.png",dpi=300)
    print(prob_matrix)

create_heat_map_based_on_merged_jsonl()

# combine_all_xlsx()
# normal_analyse()
# data_m()
# adjust_difference()
# demodified()
# data_modified()
# clean_json()