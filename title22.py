import openpyxl
import matplotlib.pyplot as plt
from collections import Counter
from openpyxl.utils import column_index_from_string
# 1. 使用 openpyxl 读取 Excel 文件并找到所有大小为 22 的标题
file_path = "Geo-AI ethics cases.xlsx"
wb = openpyxl.load_workbook(file_path)
ws = wb.active

title_cells = [cell for row in ws.iter_rows() for cell in row if cell.font and cell.font.sz == 22]

if not title_cells:
    raise ValueError("No cell with font size 22 found!")

colors = ['gray', 'red']
ii=0
for title_cell in title_cells:
    # print(title_cell.column,"ww")
    ii+=1
    # 2. 检查该标题下的列是否只有两种不同的值
    values = [ws.cell(row=row, column=title_cell.column).value for row in range(title_cell.row + 1, ws.max_row + 1) if
              ws.cell(row=row, column=title_cell.column).value is not None]
    unique_values = set(values)
    # unique_values = {value for value in values if value is not None}
    if title_cell.value=="emotional and psychological harm":
        pass
        # print(unique_values)
    if len(unique_values) == 2:
        # 3. 如果只有两种值，收集这些值并计数
        counter = Counter(values)
        # labels = list(counter.keys())
        # print(labels)
        labels = [str((label)) for label in ["False", "True"]]
        sizes = list(counter.values())
        # print(title_cell.value)
        # 4. 使用 matplotlib 绘制饼状图
        fig, ax = plt.subplots()
        wedges, texts, autotexts = ax.pie(sizes, colors=colors, autopct='%1.1f%%', startangle=90)

        # 设置图例和标题
        ax.legend(wedges, labels, title="Labels", loc="best")
        ax.set_title(f"Proportion of {title_cell.value.lower()}", fontweight='bold')
        total_count = sum(sizes)
        print(f"Proportion of {title_cell.value.lower()}")
        for label, size in zip(labels, sizes):
            proportion = (size / total_count) * 100
            print(f"{label}: {proportion:.1f}%")
        print\
            ("____________________")
        ax.axis('equal')  # Equal aspect ratio ensures that pie is drawn as a circle.
        plt.savefig(r"C:/Users/Morning/Desktop/hiwi/heart/paper/chart/"+str(title_cell.column)+".png")
        # plt.show()
